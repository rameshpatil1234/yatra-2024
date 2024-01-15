import inspect
import logging

import openpyxl
import softest


class Utils(softest.TestCase):
    def assert_list_item_text(self,list,value):
        for stop in list:
            print("The stop is:",stop.text)
            self.soft_assert(self.assertEqual,stop.text,value)
            if stop.text==value:
                print("assert pass")
            else:
                print("assert fail")
        self.assert_all()

    def custom_logger(log_Level=logging.INFO):
        logger_name=inspect.stack()[1][3]
        logger=logging.getLogger(logger_name)
        logger.setLevel(log_Level)
        fh=logging.FileHandler("automation.log")
        formatter=logging.Formatter("%(asctime)s: %(levelname)s: %(message)s",datefmt="%d/m/%Y %I-%M-%S %p")
        fh.setFormatter(formatter)
        logger.addHandler(fh)
        return logger

    def read_data_from_excel(path,sheet):
        wb=openpyxl.load_workbook(path)
        sh=wb[sheet]
        rc=sh.max_row
        cc=sh.max_column
        data_list=[]
        for i in range(2,rc+1):
            row=[]
            for j in range(1,cc+1):
                row.append(sh.cell(i,j).value)
            data_list.append(row)
        return data_list

